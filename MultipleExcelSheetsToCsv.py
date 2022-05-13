import csv

import openpyxl as xl


def xls_to_csv(file_path, csv_name):
    wb = xl.load_workbook(file_path)
    for sheet in wb.sheetnames:
        with open(f'{csv_name}_{sheet.title()}.csv', 'w') as csv_file:
            writer = csv.writer(csv_file)
            xls_sheet = wb[sheet]
            maxRow = xls_sheet.max_row+1
            maxCol = xls_sheet.max_column+1
            headers = (xls_sheet.cell(
                row=1, column=col).value for col in range(1, maxCol))
            writer.writerow(headers)
            for r in range(2, maxRow):
                xls_row = (xls_sheet.cell(
                    row=r, column=col).value for col in range(1, maxCol))
                writer.writerow(xls_row)


xls_to_csv("Identification clients GPS 2022.xlsx", "Identification clients gps")

"""
if __name__=='__main__':
import pathlib
import sys
    
try:
        file_path=sys.argv[1]
        csv_name=sys.argv[2]
except IndexError:
        sys.exit("2 arguments are required. One excel path and one save file name")
    
with pathlib.Path(file_path) as xls_file:
        if xls_file.is_file():
            xls_to_csv(xls_file,csv_name)
        else:
            sys.exit(f'Did not find {file_path}')"""


##Another way to convert only one excel sheet file to csv
#import pandas as pd
#read_file=pd.read_excel(r'Path of the file\filename.xlsx')
#read_file.to_csv(r'path to store the file.csv',index=None,header=True)
